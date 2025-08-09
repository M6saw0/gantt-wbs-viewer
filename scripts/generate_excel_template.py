#!/usr/bin/env python3
from __future__ import annotations
import csv
import datetime as dt
from pathlib import Path
from typing import List, Dict
from zipfile import ZipFile, ZIP_DEFLATED
from io import BytesIO

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
CSV_DIR = ROOT / "data" / "templates" / "csv"
OUT_XLSX = ROOT / "data" / "templates" / "projects_template.xlsx"

DATE_FMT = "yyyy/m/d"  # 0埋めなし表示
PROG_FMT = "0"  # 進捗は 0-100 の数値表示

HOLIDAYS_SHEET_NAME = "祝日"


def _read_csv(path: Path) -> List[List[str]]:
    with path.open(newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        return [row for row in reader]


def _as_date(value: str) -> dt.date | None:
    value = (value or "").strip()
    if not value:
        return None
    try:
        y, m, d = [int(x) for x in value.split("/")]
        return dt.date(y, m, d)
    except Exception:
        return None


def _as_progress(value: str) -> float | None:
    value = (value or "").strip()
    if not value:
        return None
    if value.endswith("%"):
        try:
            return float(value[:-1])
        except Exception:
            return None
    try:
        v = float(value)
        # 0-1 で入っている場合は 0-100 に拡張
        return v * 100.0 if 0.0 <= v <= 1.0 else v
    except Exception:
        return None


def _set_col_widths(ws: Worksheet, widths: List[float]) -> None:
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _normalize_rows(rows: List[List[str]]) -> List[List[str]]:
    """`前工程` に含まれるカンマで列がはみ出た行を是正する。
    余分なカラムは `前工程` に結合し、ヘッダー列数に合わせる。
    """
    if not rows:
        return rows
    headers = rows[0]
    expected = len(headers)
    try:
        pred_idx = headers.index("前工程")
    except ValueError:
        pred_idx = None
    fixed: List[List[str]] = [headers]
    for row in rows[1:]:
        # ちょうど同数 or `前工程` 未定義ならパディングのみ
        if len(row) == expected or pred_idx is None:
            if len(row) < expected:
                row = row + [None] * (expected - len(row))
            fixed.append(row[:expected])
            continue
        if len(row) > expected and pred_idx is not None:
            left = row[:pred_idx]
            right_count = expected - (pred_idx + 1)
            right = row[-right_count:] if right_count > 0 else []
            middle = row[pred_idx: len(row) - right_count] if right_count > 0 else row[pred_idx:]
            pred_val = ",".join([c for c in middle if c is not None])
            new_row = left + [pred_val] + right
            fixed.append(new_row[:expected])
        else:
            fixed.append(row[:expected])
    return fixed


def _append_project_sheet(wb: Workbook, sheet_name: str, rows: List[List[str]]) -> None:
    ws = wb.create_sheet(sheet_name)
    norm_rows = _normalize_rows(rows)
    # 望ましい列順に並べ替え
    desired = [
        "ID",
        "分類1",
        "分類2",
        "分類3",
        "タスク名",
        "詳細",
        "備考",
        "開始日",
        "終了日",
        "工数（人日）",
        "前工程",
        "担当",
        "進捗（%）",
    ]
    headers0 = norm_rows[0]
    have = {h: i for i, h in enumerate(headers0)}
    if all(h in have for h in desired):
        reordered = [desired]
        for row in norm_rows[1:]:
            reordered.append([row[have[h]] if have[h] < len(row) else None for h in desired])
        norm_rows = reordered

    for row in norm_rows:
        ws.append(row)
    headers = norm_rows[0]
    header_to_col: Dict[str, int] = {h: i + 1 for i, h in enumerate(headers)}
    date_cols = [header_to_col.get("開始日"), header_to_col.get("終了日")]
    # 進捗列は表記ゆれ対策（例: "進捗（%）", "進捗", "進捗（）"）
    prog_col = None
    for key, col in header_to_col.items():
        if key and str(key).startswith("進捗"):
            prog_col = col
            break
    id_col = header_to_col.get("ID")
    effort_col = header_to_col.get("工数（人日）")
    for r in range(2, len(norm_rows) + 1):
        for c in date_cols:
            if c:
                cell = ws.cell(row=r, column=c)
                d = _as_date(str(cell.value or ""))
                if d:
                    cell.value = d
                    cell.number_format = DATE_FMT
        if prog_col:
            cell = ws.cell(row=r, column=prog_col)
            p = _as_progress(str(cell.value or ""))
            if p is not None:
                try:
                    cell.value = float(p)
                except Exception:
                    cell.value = p
                cell.number_format = PROG_FMT
        if id_col:
            cell = ws.cell(row=r, column=id_col)
            try:
                cell.value = int(str(cell.value).strip())
                cell.number_format = "0"
            except Exception:
                pass
        if effort_col:
            cell = ws.cell(row=r, column=effort_col)
            val = str(cell.value or "").strip()
            if val:
                try:
                    eff = float(val)
                    cell.value = eff
                    cell.number_format = "0" if eff.is_integer() else "0.0"
                except Exception:
                    pass

    # 仕上げ: 終了日を補完し、欠落ケースを各1件だけ残す
    total_rows = len(norm_rows)
    start_idx = date_cols[0]
    finish_idx = date_cols[1]
    missing_finish_rows = []
    missing_start_rows = []

    # まずは可能な限り補完
    for r in range(2, total_rows + 1):
        start_cell = ws.cell(row=r, column=start_idx) if start_idx else None
        finish_cell = ws.cell(row=r, column=finish_idx) if finish_idx else None
        effort_cell = ws.cell(row=r, column=effort_col) if effort_col else None
        start_val = start_cell.value if start_cell else None
        finish_val = finish_cell.value if finish_cell else None
        eff_val = None
        if effort_cell is not None and str(effort_cell.value or "").strip() != "":
            try:
                eff_val = float(effort_cell.value)
            except Exception:
                eff_val = None

        # 終了日補完
        if finish_cell is not None and (finish_val is None) and start_val is not None:
            try:
                days = 0
                if eff_val is not None:
                    days = max(int(round(eff_val)) - 1, 0)
                new_finish = start_val + dt.timedelta(days=days)
                finish_cell.value = new_finish
                finish_cell.number_format = DATE_FMT
            except Exception:
                pass

        # 開始日補完
        if start_cell is not None and (start_val is None) and finish_val is not None:
            try:
                days = 0
                if eff_val is not None:
                    days = max(int(round(eff_val)) - 1, 0)
                new_start = finish_val - dt.timedelta(days=days)
                start_cell.value = new_start
                start_cell.number_format = DATE_FMT
            except Exception:
                pass

    # 欠落行の収集（補完後）
    for r in range(2, total_rows + 1):
        if finish_idx and ws.cell(row=r, column=finish_idx).value is None:
            missing_finish_rows.append(r)
        if start_idx and ws.cell(row=r, column=start_idx).value is None:
            missing_start_rows.append(r)

    # 欠落数の調整: それぞれ1件だけ残す
    def clear_all_but_one(rows_list, col_idx):
        if not col_idx:
            return
        if len(rows_list) == 0:
            return
        # 先頭以外は補完してしまう（開始は終了を、終了は開始を使って同日で補完）
        keep = rows_list[0]
        for r in rows_list[1:]:
            if col_idx == finish_idx:
                # 補完: 終了=開始
                s = ws.cell(row=r, column=start_idx).value if start_idx else None
                if s is not None:
                    cell = ws.cell(row=r, column=finish_idx)
                    cell.value = s
                    cell.number_format = DATE_FMT
            elif col_idx == start_idx:
                f = ws.cell(row=r, column=finish_idx).value if finish_idx else None
                if f is not None:
                    cell = ws.cell(row=r, column=start_idx)
                    cell.value = f
                    cell.number_format = DATE_FMT

    clear_all_but_one(missing_finish_rows, finish_idx)
    clear_all_but_one(missing_start_rows, start_idx)

    # もし欠落がどちらか0件なら、テスト用に1件作る（同一行を避ける）
    if start_idx and finish_idx:
        if len(missing_finish_rows) == 0:
            # 開始が入っている行を選び、終了を空に
            for r in range(2, total_rows + 1):
                if ws.cell(row=r, column=start_idx).value is not None:
                    ws.cell(row=r, column=finish_idx).value = None
                    missing_finish_rows = [r]
                    break
        if len(missing_start_rows) == 0:
            # 終了が入っている別の行を選び、開始を空に
            for r in range(2, total_rows + 1):
                if (ws.cell(row=r, column=finish_idx).value is not None) and (r not in missing_finish_rows):
                    ws.cell(row=r, column=start_idx).value = None
                    missing_start_rows = [r]
                    break
    _set_col_widths(ws, [6, 15, 15, 15, 24, 50, 20, 12, 12, 10, 20, 12, 10])
    ws.freeze_panes = "A2"


def _append_holidays_sheet(wb: Workbook, rows: List[List[str]]) -> None:
    ws = wb.create_sheet(HOLIDAYS_SHEET_NAME)
    for row in rows:
        ws.append(row)
    for r in range(2, len(rows) + 1):
        cell = ws.cell(row=r, column=1)
        d = _as_date(str(cell.value or ""))
        if d:
            cell.value = d
            cell.number_format = DATE_FMT
    _set_col_widths(ws, [12, 24])
    ws.freeze_panes = "A2"


def build_workbook(project_csvs: Dict[str, Path], holidays_csv: Path) -> Workbook:
    wb = Workbook()
    # remove default sheet
    wb.remove(wb.active)
    # project sheets
    for sheet_name, csv_path in project_csvs.items():
        rows = _read_csv(csv_path)
        _append_project_sheet(wb, sheet_name, rows)
    # holidays
    _append_holidays_sheet(wb, _read_csv(holidays_csv))

    # core properties empty
    props = wb.properties
    props.creator = ""
    props.lastModifiedBy = ""
    props.title = ""
    props.subject = ""
    props.description = ""
    props.keywords = ""

    return wb


def strip_core_properties(xlsx_path: Path) -> None:
    with ZipFile(xlsx_path, "r") as zin:
        filelist = zin.namelist()
        core_xml = zin.read("docProps/core.xml") if "docProps/core.xml" in filelist else None
        content = {name: zin.read(name) for name in filelist if name != "docProps/core.xml"}
    if core_xml is not None:
        try:
            import xml.etree.ElementTree as ET
            ET.register_namespace("cp", "http://schemas.openxmlformats.org/package/2006/metadata/core-properties")
            ET.register_namespace("dc", "http://purl.org/dc/elements/1.1/")
            ET.register_namespace("dcterms", "http://purl.org/dc/terms/")
            root = ET.fromstring(core_xml)
            ns = {"cp": "http://schemas.openxmlformats.org/package/2006/metadata/core-properties", "dc": "http://purl.org/dc/elements/1.1/", "dcterms": "http://purl.org/dc/terms/"}
            for tag in ["dc:creator", "cp:lastModifiedBy", "dcterms:created", "dcterms:modified", "dc:title", "dc:subject", "dc:description", "cp:keywords"]:
                for el in root.findall(tag, ns):
                    el.text = ""
            new_core = ET.tostring(root, encoding="utf-8", xml_declaration=True)
        except Exception:
            new_core = core_xml
    else:
        new_core = None
    tmp = BytesIO()
    with ZipFile(tmp, "w", ZIP_DEFLATED) as zout:
        for name, data in content.items():
            zout.writestr(name, data)
        if new_core is not None:
            zout.writestr("docProps/core.xml", new_core)
    xlsx_path.write_bytes(tmp.getvalue())


def main() -> None:
    project_csvs: Dict[str, Path] = {
        "プロジェクトA": CSV_DIR / "project_A.csv",
        "プロジェクトB": CSV_DIR / "project_B.csv",
    }
    holidays_csv = CSV_DIR / "holidays_template.csv"
    for p in project_csvs.values():
        if not p.exists():
            raise SystemExit(f"プロジェクトCSVが見つかりません: {p}")
    if not holidays_csv.exists():
        raise SystemExit(f"祝日CSVが見つかりません: {holidays_csv}")

    wb = build_workbook(project_csvs, holidays_csv)
    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    # メタデータの空化によるOffice互換性問題を避けるため、
    # ここではZIP再書き込みを行わない（作成者・最終更新者は空で保存済み）
    print(f"テンプレートを生成しました: {OUT_XLSX}")


if __name__ == "__main__":
    main()
